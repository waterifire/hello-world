from openpyxl import Workbook, load_workbook
import sys

wb = load_workbook('docs/mission.xlsx')
ws = wb.active

print(ws)  # prints what worksheet you are on

# ws['C2'].value = 'Dylan'  # changes the value in that cell

# ws.title = 'Data'  # changes the title

# print(ws['C2'].value)  # prints everything in the cell

working_down = True
destination_ls = ['M', 'N']
destination_ls1 = ['M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
num = '2'

# print(ws[current_cell].value)


def cell_move():
    global num
    take_out = ls[0]
    ls.remove(take_out)
    num = int(num)
    num += 1
    num = str(num)
    current_cell = 'C' + num
    ls.append(current_cell)
    print(ls)

def amount_equal():
    global num
    count = 2

    def checking_amount(count):  # example 924, 925, 926
        numb = int(num) + 1
        looking_at = 'C' + str(numb)
        if ws[ls[1]].value == ws[looking_at].value:
            check_more = True
            count += 1
            ls.append(looking_at)  # ls = (C924, C925, C926)
            while check_more:
                numb = str(int(numb) + 1)
                looking_at = 'C' + numb
                if ws[ls[-1]].value == ws[looking_at].value:
                    count += 1
                if ws[ls[-1]].value != ws[looking_at].value:
                    check_more = False
                    return count

        else:
            count = 2
            return count
    count = checking_amount(count)
    if count == 2:
        place = 'M' + num
        place2 = 'M' + str(int(num) - 1)

        val = ls[0]
        val = val[1:]
        val = "B" + val

        mal = ls[1]
        mal = mal[1:]
        mal = "B" + mal

        ws[place] = ws[val].value
        ws[place2] = ws[mal].value
    elif count > 2:
        extract_ls = []
        num = int(num)
        temp_count = 0
        extract_cell = 'B'

        for i in range(len(ls)):  # This extracts the values i need to print
            focus = ls[i]
            number_focus = focus[1:]
            extraction_value = extract_cell + number_focus
            extract_ls.append(extraction_value)

        for count1 in range(count):  # count = 3, ls =(c10, c11, c12)
            focus_row = count1 + num - 1
            temp_extra_ls = []
            temp_extra_ls.extend(extract_ls)
            item_removed = temp_extra_ls[count1]
            temp_extra_ls.remove(item_removed)
            print(f"temp===== {temp_extra_ls}")
            for i in range(len(temp_extra_ls)):
                count = 0
                for x in destination_ls:
                    drop_it = x + str(focus_row)
                    ws[drop_it] = ws[temp_extra_ls[count]].value
                    temp_count += 1
                    count += 1
        ls.pop(-1)
        num += count1 - 1

# NPQOR33)
ls = ['c1', 'c2']

while working_down:

    if ws[ls[0]].value == ws[ls[1]].value:
        print("equal!   --------------------")
        amount_equal()
        cell_move()
    else:

        print("not equal")
        cell_move()
    if ws[ls[1]].value is None:
        print("done!!!!!")
        working_down = False





wb.save('docs/mission.xlsx')

